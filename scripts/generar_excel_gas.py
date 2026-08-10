#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ACZO · Genera ACZO_Radar_Gas.xlsx desde los CSV de la carpeta. Lo llama actualizar_gas.py.
import csv, datetime, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SP = os.path.join(ROOT, "datos")
OUT = os.path.join(ROOT, "excel", "ACZO_Radar_Gas.xlsx")

A = "Calibri"
tf = Font(name=A, size=14, bold=True, color="FFFFFF")
hf = Font(name=A, size=10, bold=True, color="FFFFFF")
lf = Font(name=A, size=10, bold=True)
nf = Font(name=A, size=10)
sf = Font(name=A, size=9, italic=True, color="808080")
inf = Font(name=A, size=10, bold=True, color="0000FF")
navy = PatternFill("solid", fgColor="1F4E78")
lg = PatternFill("solid", fgColor="F2F2F2")
yel = PatternFill("solid", fgColor="FFF2CC")
grf = PatternFill("solid", fgColor="E2EFDA")
grh = PatternFill("solid", fgColor="C6E0B4")
orf = PatternFill("solid", fgColor="FCE4D6")
GRN = PatternFill("solid", fgColor="C6EFCE"); RED = PatternFill("solid", fgColor="FFC7CE")
FG = Font(name=A, size=10, color="006100"); FR = Font(name=A, size=10, color="9C0006")
thin = Side(style="thin", color="BFBFBF")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
ctr = Alignment(horizontal="center", vertical="center")
ll = Alignment(horizontal="left", vertical="center")
wctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
wl = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ---- datos ----
MEN = {}
with open(SP + "/mibgas_mensual.csv") as f:
    rd = csv.reader(f); next(rd)
    for y, m, p in rd:
        if p: MEN[(int(y), int(m))] = float(p)
DIA = []
with open(SP + "/mibgas_diario.csv") as f:
    rd = csv.reader(f); next(rd)
    for fecha, p in rd:
        DIA.append((datetime.date.fromisoformat(fecha), float(p)))
CURVA = []
if os.path.exists(SP + "/mibgas_curva.csv"):
    with open(SP + "/mibgas_curva.csv") as f:
        rd = csv.reader(f); next(rd)
        for prod, etiq, ini, fin, precio, peso in rd:
            CURVA.append((prod, etiq, float(precio), int(peso)))
FECHA_CURVA = datetime.date.today()
if os.path.exists(SP + "/mibgas_fecha.txt"):
    FECHA_CURVA = datetime.date.fromisoformat(open(SP + "/mibgas_fecha.txt").read().strip())
CFG = {"umbral_gas": 55}
if os.path.exists(os.path.join(ROOT,"config.csv")):
    with open(os.path.join(ROOT,"config.csv")) as f:
        rd = csv.reader(f); next(rd)
        for row in rd:
            if len(row) >= 2 and row[0] in CFG:
                try: CFG[row[0]] = float(row[1])
                except ValueError: pass

hoy = datetime.date.today()
NROWS = ((hoy.year + 1) - 2023 + 1) * 12
MESES = [((2023 * 12 + i) // 12, (2023 * 12 + i) % 12 + 1) for i in range(NROWS)]

wb = Workbook()

def title(ws, txt, span):
    ws.sheet_view.showGridLines = False
    t = ws.cell(row=1, column=1, value="  " + txt)
    ws.merge_cells(span); t.font = tf; t.fill = navy; t.alignment = ll
    ws.row_dimensions[1].height = 30

def head(ws, r, cols, c0=1):
    for j, h in enumerate(cols, start=c0):
        c = ws.cell(row=r, column=j, value=h); c.font = hf; c.fill = navy; c.alignment = wctr; c.border = bd
    ws.row_dimensions[r].height = 28

def marcar(ws, rng, first_cell, thr):
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f"AND(ISNUMBER({first_cell}),{first_cell}<{thr})"], fill=GRN, font=FG))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f"AND(ISNUMBER({first_cell}),{first_cell}>={thr})"], fill=RED, font=FR))

# ================= INDICE =================
wi = wb.active; wi.title = "Indice"
title(wi, "ACZO · RADAR DE GAS — MIBGAS: spot (pasado) + MIBGAS Derivatives (futuro)", "A1:H1")
wi.column_dimensions["A"].width = 3; wi.column_dimensions["B"].width = 20; wi.column_dimensions["C"].width = 100
filas = [("", "PARA QUÉ SIRVE", True),
 ("", "El equivalente en gas del Radar de Mercados eléctrico: el retrovisor (spot MIBGAS desde 2023), el mes en curso (diario, incluida la entrega de MAÑANA) y el parabrisas (curva de futuros MIBGAS Derivatives). Para valorar indexados de gas, vigilar renovaciones y controlar la plausibilidad de ofertas.", False),
 ("", "", False), ("", "HOJAS", True),
 ("MIBGAS_Spot", "Precio medio mensual del gas día-siguiente (GDAES) desde 2023 + diario del mes en curso + año en curso. Verde/rojo contra el umbral de config.csv (umbral_gas).", False),
 ("Futuros", "La curva completa: meses M+1 a M+6, trimestres Q+1 a Q+4, verano/invierno y años Y+1/Y+2, con el forward 12 meses y su control de pesos.", False),
 ("", "", False),
 ("", "ACTUALIZACIÓN: el mismo robot de las 21:15 que actualiza el radar eléctrico actualiza este archivo (carpeta _automatizacion, actualizar_gas.py). Fuente: mibgas.es, ficheros públicos MIBGAS_Data. El precio es €/MWh en PVB; el indexado de un cliente = MIBGAS del periodo + peaje variable de su grupo RL + margen de la comercializadora.", True)]
r = 3
for a, b, bold in filas:
    ca = wi.cell(row=r, column=2, value=a); ca.font = lf; ca.alignment = ll
    cb = wi.cell(row=r, column=3, value=b); cb.font = (lf if bold else nf); cb.alignment = wl
    if bold and b: cb.fill = lg; ca.fill = lg
    wi.row_dimensions[r].height = 30 if len(b) > 90 else 22
    r += 1

# ================= MIBGAS_SPOT =================
ws = wb.create_sheet("MIBGAS_Spot")
title(ws, "MIBGAS — SPOT GDAES (€/MWh): histórico mensual + mes en curso", "A1:L1")
for col, w in [("A", 13), ("B", 13), ("C", 3), ("D", 12), ("E", 12), ("F", 3), ("J", 12), ("K", 12)]:
    ws.column_dimensions[col].width = w
ws.cell(row=2, column=1, value="Izquierda: media mensual del producto día-siguiente (GDAES) por mes de ENTREGA, 2023 → hoy (las filas hasta diciembre del año próximo se rellenan solas). Centro: precios diarios del mes en curso — incluye las entregas de mañana y pasado si ya cotizaron. Derecha: año en curso. VERDE = por debajo del umbral · ROJO = por encima. El gas no tiene periodos horarios: un precio por día.").font = sf
ws.merge_cells("A2:H3"); ws.cell(row=2, column=1).alignment = wl

lb = ws.cell(row=4, column=1, value="UMBRAL de color (€/MWh) — se cambia en config.csv (umbral_gas) →")
lb.font = lf; ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=1)
lb.alignment = Alignment(horizontal="right", vertical="center")
qb = ws.cell(row=4, column=2, value=CFG["umbral_gas"]); qb.number_format = "#,##0"
qb.font = inf; qb.fill = yel; qb.border = bd; qb.alignment = ctr

head(ws, 5, ["Mes", "Media (€/MWh)"])
OD = 6
for i, (y, m) in enumerate(MESES):
    r = OD + i
    ws.cell(row=r, column=1, value=datetime.date(y, m, 1)).number_format = "mmm-yyyy"
    if (y, m) in MEN:
        ws.cell(row=r, column=2, value=MEN[(y, m)]).number_format = "#,##0.00"
    for j in (1, 2):
        cc = ws.cell(row=r, column=j); cc.font = nf; cc.border = bd; cc.alignment = ctr
OLAST = OD + NROWS - 1
marcar(ws, f"B{OD}:B{OLAST}", "$B6", "$B$4")

c = ws.cell(row=4, column=4, value=f"MES EN CURSO ({hoy:%b-%Y})"); c.font = lf; c.fill = orf
ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=5)
head(ws, 5, ["Entrega", "€/MWh"], c0=4)
r = 6
for d, p in DIA:
    ws.cell(row=r, column=4, value=d).number_format = "dd/mm"
    ws.cell(row=r, column=5, value=p).number_format = "#,##0.00"
    for j in (4, 5):
        cc = ws.cell(row=r, column=j); cc.font = nf; cc.border = bd; cc.alignment = ctr
        if d > hoy: cc.fill = orf
    r += 1
DLAST = r - 1
if DIA:
    marcar(ws, f"E6:E{DLAST}", "E6", "$B$4")
    lc = ws.cell(row=DLAST + 1, column=4, value="MEDIA MES"); lc.font = lf; lc.border = bd; lc.fill = grh
    vc = ws.cell(row=DLAST + 1, column=5, value=f"=AVERAGE(E6:E{DLAST})")
    vc.number_format = "#,##0.00"; vc.font = Font(name=A, size=11, bold=True, color="1F4E78"); vc.border = bd; vc.alignment = ctr; vc.fill = grh
    ws.cell(row=DLAST + 2, column=4, value="Filas naranjas = entregas futuras ya cotizadas (mañana, pasado).").font = sf
    ws.merge_cells(start_row=DLAST + 2, start_column=4, end_row=DLAST + 2, end_column=8)

c = ws.cell(row=4, column=10, value=f"AÑO {hoy.year}"); c.font = lf; c.fill = orf
ws.merge_cells(start_row=4, start_column=10, end_row=4, end_column=11)
head(ws, 5, ["Mes", "Media"], c0=10)
rr = 6
for i, (y, m) in enumerate(MESES):
    if y != hoy.year: continue
    ws.cell(row=rr, column=10, value=f"=A{OD + i}").number_format = "mmm"
    ws.cell(row=rr, column=11, value=f'=IF(B{OD + i}="","",B{OD + i})').number_format = "#,##0.0"
    for j in (10, 11):
        cc = ws.cell(row=rr, column=j); cc.font = nf; cc.border = bd; cc.alignment = ctr
    rr += 1
if DIA:
    ws.cell(row=rr, column=10, value=f"{hoy:%b}").font = Font(name=A, size=10, bold=True)
    ws.cell(row=rr, column=11, value=f"=E{DLAST + 1}").number_format = "#,##0.0"
    for j in (10, 11):
        cc = ws.cell(row=rr, column=j); cc.border = bd; cc.alignment = ctr; cc.fill = orf
marcar(ws, f"K6:K{rr}", "K6", "$B$4")

# ================= FUTUROS =================
wf = wb.create_sheet("Futuros")
title(wf, "MIBGAS DERIVATIVES — LA CURVA DE FUTUROS DEL GAS (€/MWh, PVB)", "A1:H1")
for col, w in [("A", 34), ("B", 13), ("C", 13), ("D", 3), ("E", 48)]:
    wf.column_dimensions[col].width = w
wf.cell(row=2, column=1, value="Lo que el mercado paga HOY por el gas de mañana: el indicador adelantado para renovaciones y para valorar precios fijos de gas. El robot toma el precio de referencia del último día de negociación de cada producto.").font = sf
wf.merge_cells("A2:E3"); wf.cell(row=2, column=1).alignment = wl
c = wf.cell(row=4, column=1, value="Fecha de la curva:"); c.font = lf
fc = wf.cell(row=4, column=2, value=FECHA_CURVA); fc.number_format = "dd/mm/yyyy"
fc.font = Font(name=A, size=10, bold=True, color="1F4E78"); fc.alignment = ctr

head(wf, 5, ["Producto (periodo de entrega)", "Precio (€/MWh)", "Peso forward"])
P0 = 6
r = P0
for prod, etiq, precio, peso in CURVA:
    lc = wf.cell(row=r, column=1, value=etiq); lc.font = nf; lc.border = bd; lc.alignment = ll
    if prod.startswith("GQES"): lc.fill = lg
    if prod.startswith(("GSES", "GYES")): lc.fill = orf
    vc = wf.cell(row=r, column=2, value=precio); vc.number_format = "#,##0.00"
    vc.font = Font(name=A, size=10, bold=True, color="1F4E78"); vc.border = bd; vc.alignment = ctr
    pc = wf.cell(row=r, column=3, value=peso); pc.number_format = "0"; pc.font = nf; pc.border = bd; pc.alignment = ctr
    r += 1
PLAST = r - 1
fr = PLAST + 2
def prow(rr, label, formula, hi=False):
    lc = wf.cell(row=rr, column=1, value=label)
    lc.font = (Font(name=A, size=11, bold=True, color="215C2E") if hi else lf); lc.border = bd; lc.alignment = ll
    lc.fill = (grh if hi else grf)
    vc = wf.cell(row=rr, column=2, value=formula); vc.number_format = "#,##0.0"; vc.border = bd; vc.alignment = ctr
    vc.font = (Font(name=A, size=11, bold=True, color="1F4E78") if hi else lf); vc.fill = (grh if hi else grf)
prow(fr, "Suma de pesos (debe ser 12)", f"=SUM(C{P0}:C{PLAST})")
wf.cell(row=fr, column=3, value=f'=IF(B{fr}=12,"OK","⚠")').font = Font(name=A, size=10, bold=True)
prow(fr + 1, "FORWARD PRÓXIMOS 12 MESES (€/MWh)", f"=SUMPRODUCT(B{P0}:B{PLAST},C{P0}:C{PLAST})/SUM(C{P0}:C{PLAST})", hi=True)
prow(fr + 2, "Spot del mes en curso (€/MWh)", "=MIBGAS_Spot!E" + str(DLAST + 1 if DIA else 6))
prow(fr + 3, "Contango/backwardation (forward − spot)", f"=B{fr + 1}-B{fr + 2}")

NOTAS = [
("CÓMO SE LEE", 1),
("· Forward > spot (contango): el mercado espera gas más caro — renovar pronto y revisar fijaciones.", 0),
("· Forward < spot (backwardation): el mercado espera relajación — favorece indexado y esperar.", 0),
("· Estacionalidad estructural: el invierno (GSES_W) cotiza por encima del verano (GSES_S). Un fijo anual promedia las dos.", 0),
("· Control de plausibilidad de ofertas (ficha de gas): energía sin peajes esperada ≈ forward + 15-25 €/MWh de margen.", 0),
("· El precio TUR de materia prima se referencia a MIBGAS: esta curva ANTICIPA los reprecios trimestrales de la TUR (ene/abr/jul/oct).", 0),
]
nr = 5
for txt, b in NOTAS:
    c = wf.cell(row=nr, column=5, value=txt); c.font = (lf if b else Font(name=A, size=9)); c.alignment = wl
    if b: c.fill = lg
    wf.row_dimensions[nr].height = 24
    nr += 1

import os as _os
tmp = OUT + ".tmp"
wb.save(tmp)
_os.replace(tmp, OUT)
print("Guardado:", OUT, "| meses:", len(MEN), "| dias mes:", len(DIA), "| curva:", len(CURVA))
