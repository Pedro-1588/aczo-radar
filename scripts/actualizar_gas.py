#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ACZO · Actualizador del Radar de Gas (MIBGAS).
1) Descarga los ficheros anuales MIBGAS_Data (2023→hoy): spot GDAES_D+1 diario.
2) Construye el mensual histórico y el diario del mes en curso.
3) Lee la curva de futuros MIBGAS Derivatives del XLSX (M+1..M+6, Q+1..Q+4, W/S, Y+1/Y+2)
   y calcula los pesos del forward 12 meses.
4) Reconstruye ACZO_Radar_Gas.xlsx.
Lo invoca actualizar_mercados.py cada día a las 21:15 (launchd). Sin argumentos.
"""
import csv, datetime, os, ssl, sys, subprocess
import urllib.request

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPTS = os.path.dirname(os.path.abspath(__file__))
SP = os.path.join(ROOT, "datos")
LOG = os.path.join(ROOT, "log.txt")
CTX = ssl.create_default_context(); CTX.check_hostname = False; CTX.verify_mode = ssl.CERT_NONE

def log(msg):
    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    with open(LOG, "a") as f: f.write(f"{stamp}  {msg}\n")
    print(msg)

def bajar(url, destino, binario=False):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, context=CTX, timeout=60) as r:
        datos = r.read()
    if not binario and b"<!DOCTYPE" in datos[:200]:
        raise RuntimeError(f"MIBGAS devolvió HTML en vez de datos: {url}")
    with open(destino, "wb") as f: f.write(datos)
    return len(datos)

def fecha_es(s):
    d, m, y = s.strip().split("/")
    return datetime.date(int(y), int(m), int(d))

def leer_spot_csv(path):
    """Devuelve {fecha_entrega: precio} del producto GDAES_D+n (n mínimo por fecha)."""
    spot = {}
    with open(path, encoding="utf-8", errors="replace") as f:
        for ln in f:
            campos = [c.strip().strip('"') for c in ln.split(";")]
            if len(campos) < 7 or not campos[1].startswith("GDAES_D+"): continue
            try:
                entrega = fecha_es(campos[4]); precio = float(campos[6]); n = int(campos[1].split("+")[1])
            except (ValueError, IndexError):
                continue
            if entrega not in spot or n < spot[entrega][0]:
                spot[entrega] = (n, precio)
    return {d: p for d, (n, p) in spot.items()}

def actualizar_spot():
    hoy = datetime.date.today()
    spot = {}
    for y in range(2023, hoy.year + 1):
        destino = os.path.join(SP, f"mibgas_{y}.csv")
        if y == hoy.year or not os.path.exists(destino):
            bajar(f"https://www.mibgas.es/es/file-access/MIBGAS_Data_{y}.csv?path=AGNO_{y}/XLS", destino)
        spot.update(leer_spot_csv(destino))
    log(f"MIBGAS: spot con {len(spot)} días de entrega (último {max(spot)})")
    # mensual (meses cerrados)
    with open(os.path.join(SP, "mibgas_mensual.csv"), "w", newline="") as f:
        w = csv.writer(f); w.writerow(["y", "m", "precio"])
        y_, m_ = 2023, 1
        while (y_, m_) < (hoy.year, hoy.month):
            vals = [p for d, p in spot.items() if (d.year, d.month) == (y_, m_)]
            w.writerow([y_, m_, f"{sum(vals)/len(vals):.2f}" if vals else ""])
            y_, m_ = (y_, m_ + 1) if m_ < 12 else (y_ + 1, 1)
    # diario del mes en curso (incluye entregas futuras ya cotizadas: mañana, pasado…)
    with open(os.path.join(SP, "mibgas_diario.csv"), "w", newline="") as f:
        w = csv.writer(f); w.writerow(["fecha", "precio"])
        for d in sorted(d for d in spot if (d.year, d.month) == (hoy.year, hoy.month)):
            w.writerow([d.isoformat(), f"{spot[d]:.2f}"])
    return spot

def actualizar_curva():
    from openpyxl import load_workbook
    hoy = datetime.date.today()
    destino = os.path.join(SP, "mibgas_curva.xlsx")
    bajar(f"https://www.mibgas.es/es/file-access/MIBGAS_Data_{hoy.year}.xlsx?path=AGNO_{hoy.year}/XLS",
          destino, binario=True)
    wb = load_workbook(destino, read_only=True)
    ws = wb["Trading Data PVB&VTP"]
    PRODUCTOS = ["GMAES", "GMES_M+2", "GMES_M+3", "GMES_M+4", "GMES_M+5", "GMES_M+6",
                 "GQES_Q+1", "GQES_Q+2", "GQES_Q+3", "GQES_Q+4",
                 "GSES_S", "GSES_W", "GYES_Y+1", "GYES_Y+2"]
    ultimo = {}   # producto -> (trading_day, first, last, precio)
    for r in ws.iter_rows(min_row=2, max_col=9, values_only=True):
        td, prod, _, area = r[0], r[1], r[2], r[3]
        if prod not in PRODUCTOS or area != "ES" or td is None: continue
        precio = next((v for v in (r[6], r[8], r[7]) if v is not None), None)  # ref > last > auction
        if precio is None: continue
        if prod not in ultimo or td > ultimo[prod][0]:
            ultimo[prod] = (td, r[4].date(), r[5].date(), float(precio))
    if "GMAES" not in ultimo or "GQES_Q+1" not in ultimo:
        raise RuntimeError("MIBGAS: curva sin productos M/Q — se mantiene la última guardada")
    fecha_curva = max(v[0] for v in ultimo.values()).date()
    # cobertura mensual de cada producto
    def meses(first, last):
        out = []; y, m = first.year, first.month
        while (y, m) <= (last.year, last.month):
            out.append((y, m)); y, m = (y, m + 1) if m < 12 else (y + 1, 1)
        return out
    cobertura = {p: meses(v[1], v[2]) for p, v in ultimo.items()}
    # pesos del forward: 12 meses desde el mes que viene
    y, m = (fecha_curva.year, fecha_curva.month + 1) if fecha_curva.month < 12 else (fecha_curva.year + 1, 1)
    objetivo = [((y * 12 + m - 1 + i) // 12, (y * 12 + m - 1 + i) % 12 + 1) for i in range(12)]
    q_prods = [p for p in ultimo if p.startswith("GQES")]
    m_prods = {cobertura[p][0]: p for p in ultimo if p.startswith(("GMAES", "GMES"))}
    pesos = {}
    sin_cubrir = 0
    for mes in objetivo:
        q = next((p for p in q_prods if mes in cobertura[p]), None)
        if q and all(mm in objetivo for mm in cobertura[q]):
            pesos[q] = pesos.get(q, 0) + 1
        elif mes in m_prods:
            pesos[m_prods[mes]] = pesos.get(m_prods[mes], 0) + 1
        elif q:
            pesos[q] = pesos.get(q, 0) + 1
        else:
            an = next((p for p in ("GYES_Y+1", "GYES_Y+2") if p in ultimo and mes in cobertura[p]), None)
            if an: pesos[an] = pesos.get(an, 0) + 1
            else: sin_cubrir += 1
    if sin_cubrir: log(f"MIBGAS: {sin_cubrir} meses del forward sin producto (chivato en el Excel)")
    ETIQ = {"GMAES": "Mes M+1", "GMES_M+2": "Mes M+2", "GMES_M+3": "Mes M+3", "GMES_M+4": "Mes M+4",
            "GMES_M+5": "Mes M+5", "GMES_M+6": "Mes M+6", "GQES_Q+1": "Trimestre Q+1", "GQES_Q+2": "Trimestre Q+2",
            "GQES_Q+3": "Trimestre Q+3", "GQES_Q+4": "Trimestre Q+4", "GSES_S": "Verano (abr-sep)",
            "GSES_W": "Invierno (oct-mar)", "GYES_Y+1": "Año Y+1", "GYES_Y+2": "Año Y+2"}
    with open(os.path.join(SP, "mibgas_curva.csv"), "w", newline="") as f:
        w = csv.writer(f); w.writerow(["producto", "etiqueta", "entrega_ini", "entrega_fin", "precio", "peso"])
        for p in PRODUCTOS:
            if p not in ultimo: continue
            _, ini, fin, precio = ultimo[p]
            w.writerow([p, f"{ETIQ[p]} ({ini:%b%y}-{fin:%b%y})", ini.isoformat(), fin.isoformat(),
                        f"{precio:.2f}", pesos.get(p, 0)])
    with open(os.path.join(SP, "mibgas_fecha.txt"), "w") as f:
        f.write(fecha_curva.isoformat())
    log(f"MIBGAS: curva de {len(ultimo)} productos ({fecha_curva}) · pesos suman {sum(pesos.values())}")

if __name__ == "__main__":
    try:
        actualizar_spot()
    except Exception as e:
        log(f"ERROR MIBGAS spot: {e} — el Excel se regenera con lo último disponible")
    try:
        actualizar_curva()
    except Exception as e:
        log(f"ERROR MIBGAS curva: {e} — el Excel mantiene la última curva guardada")
    r = subprocess.run([sys.executable, os.path.join(SCRIPTS,"generar_excel_gas.py")],
                       capture_output=True, text=True)
    if r.returncode == 0:
        log("Excel gas regenerado OK")
    else:
        log(f"ERROR generando Excel gas: {r.stderr.strip()[-400:]}")
        sys.exit(1)
