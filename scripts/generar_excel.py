import csv, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

import os
ROOT=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SP=os.path.join(ROOT,"datos")
OUT=os.path.join(ROOT,"excel","ACZO_Radar_Mercados.xlsx")

A="Calibri"
tf=Font(name=A,size=14,bold=True,color="FFFFFF")
hf=Font(name=A,size=10,bold=True,color="FFFFFF")
lf=Font(name=A,size=10,bold=True)
nf=Font(name=A,size=10)
sf=Font(name=A,size=9,italic=True,color="808080")
inf=Font(name=A,size=10,bold=True,color="0000FF")
navy=PatternFill("solid",fgColor="1F4E78")
lg=PatternFill("solid",fgColor="F2F2F2")
yel=PatternFill("solid",fgColor="FFF2CC")
grf=PatternFill("solid",fgColor="E2EFDA")
grh=PatternFill("solid",fgColor="C6E0B4")
orf=PatternFill("solid",fgColor="FCE4D6")
gry=PatternFill("solid",fgColor="D9D9D9")
GRN=PatternFill("solid",fgColor="C6EFCE"); RED=PatternFill("solid",fgColor="FFC7CE")
FG=Font(name=A,size=10,color="006100"); FR=Font(name=A,size=10,color="9C0006")
thin=Side(style="thin",color="BFBFBF")
bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center",vertical="center")
ll=Alignment(horizontal="left",vertical="center")
wctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
wl=Alignment(horizontal="left",vertical="center",wrap_text=True)

# ---------------- datos ----------------
OM=[]
with open(SP+"/omie_mensual_periodos.csv") as f:
    rd=csv.reader(f); next(rd)
    for y,m,p1,p2,p3 in rd: OM.append((int(y),int(m),float(p1),float(p2),float(p3)))
OM6=[]
with open(SP+"/omie_6p.csv") as f:
    rd=csv.reader(f); next(rd)
    for row in rd:
        y,m=int(row[0]),int(row[1])
        OM6.append((y,m,[float(v) if v else None for v in row[2:8]]))
DAILY=[]
with open(SP+"/omie_diario_minmax.csv") as f:
    rd=csv.reader(f); next(rd)
    for fecha,mn,me,mx in rd:
        DAILY.append((datetime.date.fromisoformat(fecha),float(mn),float(me),float(mx)))

CFG={"umbral_20":110,"umbral_6p":110,"mejor_fijo":109.9,"peajes":28,"cobertura":10}
if os.path.exists(os.path.join(ROOT,"config.csv")):
    with open(os.path.join(ROOT,"config.csv")) as f:
        rd=csv.reader(f); next(rd)
        for row in rd:
            if len(row)>=2 and row[0] in CFG:
                try: CFG[row[0]]=float(row[1])
                except ValueError: pass

FECHA_OMIP=datetime.date(2026,8,7)
if os.path.exists(SP+"/omip.csv"):
    OMIP_CSV=[]
    with open(SP+"/omip.csv") as f:
        rd=csv.reader(f); next(rd)
        for k,name,price,peso in rd:
            OMIP_CSV.append((k,name,float(price),int(peso)))
    if os.path.exists(SP+"/omip_fecha.txt"):
        FECHA_OMIP=datetime.date.fromisoformat(open(SP+"/omip_fecha.txt").read().strip())
else:
    OMIP_CSV=None

def addm(y,m,k):
    t=(y*12+m-1)+k
    return t//12, t%12+1

_hoy=datetime.date.today()
NROWS=((_hoy.year+1)-2023+1)*12  # ene-2023 -> dic del año próximo
MESES=[addm(2023,1,i) for i in range(NROWS)]
NFROZ=len(OM)   # 42
# temporadas 3.0/6.1: mes -> (periodo alto, periodo bajo); P6 siempre
SEASON={1:(1,2),2:(1,2),3:(2,3),4:(4,5),5:(4,5),6:(3,4),7:(1,2),8:(3,4),9:(3,4),10:(4,5),11:(2,3),12:(1,2)}
SNAME={1:"ALTA",2:"ALTA",3:"M-ALTA",4:"BAJA",5:"BAJA",6:"MEDIA",7:"ALTA",8:"MEDIA",9:"MEDIA",10:"BAJA",11:"M-ALTA",12:"ALTA"}

wb=Workbook()

def title(ws,txt,span="A1:K1"):
    ws.sheet_view.showGridLines=False
    t=ws.cell(row=1,column=1,value="  "+txt)
    ws.merge_cells(span); t.font=tf; t.fill=navy; t.alignment=ll; ws.row_dimensions[1].height=30

def head(ws,r,cols,c0=1):
    for j,h in enumerate(cols,start=c0):
        c=ws.cell(row=r,column=j,value=h); c.font=hf; c.fill=navy; c.alignment=wctr; c.border=bd
    ws.row_dimensions[r].height=28

def marcar(ws,rng,first_cell,thr):
    ws.conditional_formatting.add(rng,FormulaRule(formula=[f"AND(ISNUMBER({first_cell}),{first_cell}<{thr})"],fill=GRN,font=FG))
    ws.conditional_formatting.add(rng,FormulaRule(formula=[f"AND(ISNUMBER({first_cell}),{first_cell}>={thr})"],fill=RED,font=FR))

# ================================================== INDICE
wi=wb.active; wi.title="Indice"
title(wi,"ACZO · RADAR DE MERCADOS — OMIE (pasado) + OMIP (futuro) · 2.0TD y 3.0/6.1TD","A1:H1")
wi.column_dimensions["A"].width=3; wi.column_dimensions["B"].width=20
wi.column_dimensions["C"].width=100
rows=[("","PARA QUÉ SIRVE",True),
 ("","Panel de seguimiento de mercado para el equipo: el retrovisor (spot OMIE desde 2023), el mes en curso (diario) y el parabrisas (futuros OMIP). Sin cliente cargado: es la foto del mercado. Los estudios de cliente viven en los otros Excel.",False),
 ("","",False),("","HOJAS",True),
 ("OMIE_20","Spot mensual por periodo P1/P2/P3 (calendario 2.0TD) + precios diarios del mes en curso + año en curso. Verde/rojo contra el umbral que fijes.",False),
 ("OMIE_6P","Spot mensual por periodo P1…P6 (calendario 3.0TD/6.1TD, con temporadas). Cada mes solo cotiza los periodos de SU temporada, más P6.",False),
 ("OMIP","Futuros: curva mensual, trimestral y anual + forward 12 meses + ventana de fijación. Con consulta automática o manual (celdas amarillas).",False),
 ("PQ_*","Zonas Power Query OPCIONALES (Guia_PowerQuery_Mercados.md). Con el actualizador automático instalado NO hacen falta: los datos llegan ya escritos.",False),
 ("","",False),
 ("","ACTUALIZACIÓN: el actualizador automático (carpeta _automatizacion) reescribe este archivo cada día a las 21:15 con OMIE (incluido el precio de MAÑANA, publicado ~13:00) y la curva OMIP del día. Tú solo abres el Excel. Si el Mac estaba apagado a esa hora, se ejecuta al encenderlo. No dejes el archivo abierto permanentemente: al estar abierto, la actualización de ese día no puede escribirse.",True)]
r=3
for a,b,bold in rows:
    ca=wi.cell(row=r,column=2,value=a); ca.font=lf; ca.alignment=ll
    cb=wi.cell(row=r,column=3,value=b); cb.font=(lf if bold else nf); cb.alignment=wl
    if bold and b: cb.fill=lg; ca.fill=lg
    wi.row_dimensions[r].height=30 if len(b)>90 else 22
    r+=1

# ================================================== OMIE_20
wo=wb.create_sheet("OMIE_20")
title(wo,"OMIE — SPOT 2.0TD: histórico mensual + mes en curso","A1:R1")
for col,w in [("A",13),("B",12),("C",12),("D",12),("E",12),("F",13),("G",3),
              ("H",11),("I",11),("J",12),("K",11),("L",3),("Q",12),("R",12)]:
    wo.column_dimensions[col].width=w
wo.cell(row=2,column=1,value="Izquierda: spot medio mensual por periodo 2.0TD (2023 → hoy; las filas hasta dic-2027 se rellenan solas con Power Query al cerrar cada mes). Centro: precios diarios del mes en curso (mín/media/máx cuartohorarios). Derecha: año en curso. VERDE = mes por debajo del umbral · ROJO = por encima. Fuente: omie.es marginalpdbc (validado 0,0% vs informe oficial OMIE).").font=sf
wo.merge_cells("A2:G3"); wo.cell(row=2,column=1).alignment=wl

lb=wo.cell(row=4,column=1,value="UMBRAL de color (€/MWh) — p.ej. mejor fijo del catálogo o quiebre del cliente →")
lb.font=lf; wo.merge_cells(start_row=4,start_column=1,end_row=4,end_column=5)
lb.alignment=Alignment(horizontal="right",vertical="center")
qb=wo.cell(row=4,column=6,value=CFG["umbral_20"]); qb.number_format="#,##0"
qb.font=inf; qb.fill=yel; qb.border=bd; qb.alignment=ctr

head(wo,5,["Mes","P1 punta","P2 llano","P3 valle","Media (€/kWh)","Media (€/MWh)"])
OD=6
for i,(y,m) in enumerate(MESES):
    r=OD+i
    wo.cell(row=r,column=1,value=datetime.date(y,m,1)).number_format="mmm-yyyy"
    froz=i<NFROZ
    for j,st in ((2,"S"),(3,"T"),(4,"U")):
        fb=f"{st}{r}" if froz else '""'
        wo.cell(row=r,column=j,value=f"=IFERROR(VLOOKUP($A{r},PQ_Mensual!$A:$D,{j},0),{fb})").number_format="#,##0.00000"
    if froz:
        for j,v in ((19,OM[i][2]),(20,OM[i][3]),(21,OM[i][4])):
            wo.cell(row=r,column=j,value=v).number_format="#,##0.00000"
    wo.cell(row=r,column=5,value=f'=IF(COUNT(B{r}:D{r})=0,"",AVERAGE(B{r}:D{r}))').number_format="#,##0.00000"
    wo.cell(row=r,column=6,value=f'=IF(E{r}="","",E{r}*1000)').number_format="#,##0.0"
    for j in range(1,7):
        cc=wo.cell(row=r,column=j); cc.font=nf; cc.border=bd; cc.alignment=ctr
OLAST=OD+NROWS-1
marcar(wo,f"F{OD}:F{OLAST}","$F6","$F$4")

# ---- mes en curso (diario, PQ_Diario con congelado en V:X) ----
c=wo.cell(row=3,column=8,value="Mes del panel →"); c.font=sf
mc=wo.cell(row=3,column=9,value=f"=IFERROR(DATE(YEAR(MAX(PQ_Diario!A:A)),MONTH(MAX(PQ_Diario!A:A)),1),DATE({DAILY[0][0].year},{DAILY[0][0].month},1))")
mc.number_format="mmm-yyyy"; mc.font=Font(name=A,size=10,bold=True,color="1F4E78"); mc.alignment=ctr
c=wo.cell(row=4,column=8,value="MES EN CURSO — precios diarios (€/MWh, cuartohorario)")
c.font=lf; c.fill=orf
wo.merge_cells(start_row=4,start_column=8,end_row=4,end_column=11)
head(wo,5,["Día","Mínimo","Media","Máximo"],c0=8)
DD=6
STAT={d.day:(mn,me,mx) for d,mn,me,mx in DAILY}
for i in range(31):
    r=DD+i; n=i+1
    wo.cell(row=r,column=8,value=f'=IF({n}>DAY(EOMONTH($I$3,0)),"",DATE(YEAR($I$3),MONTH($I$3),{n}))').number_format="dd/mm"
    for j,st in ((9,22),(10,23),(11,24)):
        fb=f"{get_column_letter(st)}{r}" if n in STAT else '""'
        wo.cell(row=r,column=j,value=f'=IF($H{r}="","",IFERROR(VLOOKUP($H{r},PQ_Diario!$A:$D,{j-7},0),{fb}))').number_format="#,##0.00"
    if n in STAT:
        for j,v in zip((22,23,24),STAT[n]):
            wo.cell(row=r,column=j,value=v).number_format="#,##0.00"
    for j in (8,9,10,11):
        cc=wo.cell(row=r,column=j); cc.font=nf; cc.border=bd; cc.alignment=ctr
DLAST=DD+30
for col in ("I","J","K"):
    marcar(wo,f"{col}{DD}:{col}{DLAST}",f"{col}6","$F$4")
sr=DLAST+1
def qrow(r,label,rng_ini,rng_fin,hi=False):
    lc=wo.cell(row=r,column=8,value=label); lc.font=lf; lc.border=bd; lc.alignment=ll
    for j,col in ((9,"I"),(10,"J"),(11,"K")):
        vc=wo.cell(row=r,column=j,value=f"=IFERROR(AVERAGE({col}{rng_ini}:{col}{rng_fin}),0)")
        vc.number_format="#,##0.00"; vc.border=bd; vc.alignment=ctr
        vc.font=(Font(name=A,size=11,bold=True,color="1F4E78") if hi else lf)
        vc.fill=(grf if hi else lg)
    wo.cell(row=r,column=8).fill=(grf if hi else lg)
qrow(sr,  "1ª quincena",DD,DD+14)
qrow(sr+1,"2ª quincena",DD+15,DLAST)
qrow(sr+2,"MEDIA MES",DD,DLAST,hi=True)
wo.cell(row=sr+3,column=8,value="vs media últimos 12 meses cerrados").font=sf
wo.merge_cells(start_row=sr+3,start_column=8,end_row=sr+3,end_column=9)
vc=wo.cell(row=sr+3,column=10,value=f'=IFERROR(J{sr+2}/AVERAGEIF(F{OD}:F{OLAST},">0")-1,0)'); vc.number_format="+0%;-0%"
vc.font=Font(name=A,size=10,bold=True); vc.alignment=ctr; vc.border=bd
nn=wo.cell(row=sr+4,column=8,value="ORIGEN: si PQ_Mensual / PQ_Diario tienen datos (Power Query), mandan ellas; si no, valores congelados (columnas ocultas). Ver 'Guia_PowerQuery_Mercados.md'.")
nn.font=sf; wo.merge_cells(start_row=sr+4,start_column=8,end_row=sr+5,end_column=15); nn.alignment=wl
for col in "STUVWX": wo.column_dimensions[col].hidden=True

# ---- año en curso ----
yr=DAILY[0][0].year
c=wo.cell(row=4,column=17,value=f"AÑO {yr} (€/MWh)"); c.font=lf; c.fill=orf
wo.merge_cells(start_row=4,start_column=17,end_row=4,end_column=18)
head(wo,5,["Mes","Media"],c0=17)
rr=6
for i,(y,m) in enumerate(MESES):
    if y!=yr: continue
    wo.cell(row=rr,column=17,value=f"=A{OD+i}").number_format="mmm"
    wo.cell(row=rr,column=18,value=f'=IF(F{OD+i}="","",F{OD+i})').number_format="#,##0.0"
    for j in (17,18):
        cc=wo.cell(row=rr,column=j); cc.font=nf; cc.border=bd; cc.alignment=ctr
    rr+=1
wo.cell(row=rr,column=17,value="=$I$3").number_format="mmm"
wo.cell(row=rr,column=18,value=f"=J{sr+2}"); wo.cell(row=rr,column=18).number_format="#,##0.0"
for j in (17,18):
    cc=wo.cell(row=rr,column=j); cc.font=Font(name=A,size=10,bold=True); cc.border=bd; cc.alignment=ctr; cc.fill=orf
marcar(wo,f"R6:R{rr}","R6","$F$4")

# ================================================== OMIE_6P
w6=wb.create_sheet("OMIE_6P")
title(w6,"OMIE — SPOT 3.0TD / 6.1TD: mensual por periodo P1…P6 (con temporadas)","A1:L1")
for col,w in [("A",13),("B",11),("C",11),("D",11),("E",11),("F",11),("G",11),("H",13),("I",10)]:
    w6.column_dimensions[col].width=w
w6.cell(row=2,column=1,value="Mismo spot OMIE clasificado con el calendario 6 periodos (Circular 3/2020): cada mes solo cotiza los DOS periodos de su temporada + P6 (0-8h laborables, findes y festivos). Temporadas: ALTA ene/feb/jul/dic (P1-P2) · MEDIA-ALTA mar/nov (P2-P3) · MEDIA jun/ago/sep (P3-P4) · BAJA abr/may/oct (P4-P5). Las celdas grises no cotizan ese mes. Base para valorar indexados 3.0/6.1 (pool + fee) y la jugada de cierres trimestrales.").font=sf
w6.merge_cells("A2:I3"); w6.cell(row=2,column=1).alignment=wl

lb=w6.cell(row=4,column=1,value="UMBRAL de color (€/MWh) →"); lb.font=lf
w6.merge_cells(start_row=4,start_column=1,end_row=4,end_column=6)
lb.alignment=Alignment(horizontal="right",vertical="center")
qb=w6.cell(row=4,column=8,value=CFG["umbral_6p"]); qb.number_format="#,##0"
qb.font=inf; qb.fill=yel; qb.border=bd; qb.alignment=ctr

head(w6,5,["Mes","P1","P2","P3","P4","P5","P6","Media (€/MWh)","Temporada"])
for i,(y,m) in enumerate(MESES):
    r=OD+i
    w6.cell(row=r,column=1,value=datetime.date(y,m,1)).number_format="mmm-yyyy"
    hi,lo=SEASON[m]; act={hi,lo,6}
    froz=i<NFROZ
    for p in range(1,7):
        j=p+1
        cc=w6.cell(row=r,column=j)
        if p in act:
            if froz:
                v6=OM6[i][2][p-1]
                fb=f"{get_column_letter(18+p)}{r}"
                w6.cell(row=r,column=18+p,value=v6).number_format="#,##0.00000"
            else:
                fb='""'
            cc.value=f"=IFERROR(VLOOKUP($A{r},PQ_Mensual6P!$A:$G,{j},0),{fb})"
            cc.number_format="#,##0.00000"
        else:
            cc.fill=gry
        cc.font=nf; cc.border=bd; cc.alignment=ctr
    w6.cell(row=r,column=8,value=f'=IF(COUNT(B{r}:G{r})=0,"",AVERAGE(B{r}:G{r})*1000)').number_format="#,##0.0"
    w6.cell(row=r,column=9,value=SNAME[m])
    for j in (1,8,9):
        cc=w6.cell(row=r,column=j); cc.font=nf; cc.border=bd; cc.alignment=ctr
marcar(w6,f"H{OD}:H{OLAST}","$H6","$H$4")
nn=w6.cell(row=OLAST+2,column=1,value="Media = promedio simple de los periodos que cotizan ese mes (indicativa; el coste real de un cliente depende de su mix por periodos). ORIGEN: PQ_Mensual6P si tiene datos; si no, congelado (columnas ocultas S:X). La media ponderada por consumo se calcula en el Excel 3.0 con la curva del cliente.")
nn.font=sf; w6.merge_cells(start_row=OLAST+2,start_column=1,end_row=OLAST+3,end_column=9); nn.alignment=wl
for col in "STUVWX": w6.column_dimensions[col].hidden=True

# ================================================== OMIP
wp=wb.create_sheet("OMIP")
title(wp,"OMIP — FUTUROS ELÉCTRICOS ESPAÑA (FTB): curva forward + ventana de fijación","A1:J1")
for col,w in [("A",8),("B",16),("C",13),("D",13),("E",13),("F",3),("G",42),("H",13)]:
    wp.column_dimensions[col].width=w
wp.cell(row=2,column=1,value="El parabrisas: lo que el mercado paga HOY por la electricidad de mañana. Los fijos de las comercializadoras se construyen sobre esta curva (forward + peajes/ajustes + cobertura + fee), así que se mueve ANTES que el spot: es el indicador adelantado del radar. Con la consulta PQ_OMIP se actualiza sola; si no, teclea la columna Manual (amarilla, 5 min/semana desde omip.pt).").font=sf
wp.merge_cells("A2:H3"); wp.cell(row=2,column=1).alignment=wl

c=wp.cell(row=4,column=1,value="Fecha de la lectura:"); c.font=lf
wp.merge_cells(start_row=4,start_column=1,end_row=4,end_column=2)
fc=wp.cell(row=4,column=3,value=FECHA_OMIP); fc.number_format="dd/mm/yyyy"
fc.font=inf; fc.fill=yel; fc.border=bd; fc.alignment=ctr
c=wp.cell(row=4,column=4,value="(con PQ_OMIP cargado, la lectura es la del último refresco)"); c.font=sf
wp.merge_cells(start_row=4,start_column=4,end_row=4,end_column=8)

head(wp,5,["Clave","Instrumento","Precio (€/MWh)","Manual","Peso forward (meses)"])
OMIP_FROZ=[("M1","FTB M Sep-26",102.50,1),("M2","FTB M Oct-26",97.00,1),
 ("M3","FTB M Nov-26",99.35,1),("M4","FTB M Dec-26",102.00,1),
 ("M5","FTB M Jan-27",96.56,0),("M6","FTB M Feb-27",84.90,0),
 ("Q1","FTB Q4-26",99.45,0),("Q2","FTB Q1-27",83.50,3),("Q3","FTB Q2-27",45.85,3),
 ("Q4","FTB Q3-27",67.75,2),("Q5","FTB Q4-27",68.66,0),("Q6","FTB Q1-28",62.11,0),
 ("Q7","FTB Q2-28",39.05,0),
 ("Y1","FTB YR-27",66.40,0),("Y2","FTB YR-28",55.00,0),("Y3","FTB YR-29",52.60,0)]
if OMIP_CSV: OMIP_FROZ=OMIP_CSV

P0=6
for i,(k,name,price,peso) in enumerate(OMIP_FROZ):
    r=P0+i
    kc=wp.cell(row=r,column=1,value=k); kc.font=lf; kc.border=bd; kc.alignment=ctr; kc.fill=lg
    nc=wp.cell(row=r,column=2,value=f'=IFERROR(VLOOKUP($A{r},PQ_OMIP!$A:$C,2,0),J{r})'); nc.font=nf; nc.border=bd; nc.alignment=ll
    pc=wp.cell(row=r,column=3,value=f'=IFERROR(VLOOKUP($A{r},PQ_OMIP!$A:$C,3,0),D{r})')
    pc.number_format="#,##0.00"; pc.font=Font(name=A,size=10,bold=True,color="1F4E78"); pc.border=bd; pc.alignment=ctr
    mc=wp.cell(row=r,column=4,value=price); mc.number_format="#,##0.00"; mc.font=inf; mc.fill=yel; mc.border=bd; mc.alignment=ctr
    wc_=wp.cell(row=r,column=5,value=peso); wc_.number_format="0"; wc_.font=inf; wc_.fill=yel; wc_.border=bd; wc_.alignment=ctr
    wp.cell(row=r,column=10,value=name)  # nombre congelado (oculto)
PLAST=P0+len(OMIP_FROZ)-1
wp.column_dimensions["J"].hidden=True

fr=PLAST+2
def prow(r,label,formula,fmt="#,##0.0",hi=False,yellow=None):
    lc=wp.cell(row=r,column=1,value=label)
    lc.font=(Font(name=A,size=11,bold=True,color="215C2E") if hi else lf); lc.border=bd; lc.alignment=ll
    wp.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    lc.fill=(grh if hi else grf)
    vc=wp.cell(row=r,column=3,value=formula if yellow is None else yellow)
    vc.number_format=fmt; vc.border=bd; vc.alignment=ctr
    if yellow is not None: vc.font=inf; vc.fill=yel
    else:
        vc.font=(Font(name=A,size=11,bold=True,color="1F4E78") if hi else lf); vc.fill=(grh if hi else grf)
prow(fr,"Suma de pesos (debe ser 12)",f"=SUM(E{P0}:E{PLAST})","0")
wc_=wp.cell(row=fr,column=4,value=f'=IF(C{fr}=12,"OK","⚠ ajusta pesos")'); wc_.font=Font(name=A,size=10,bold=True); wc_.alignment=ll
prow(fr+1,"FORWARD PRÓXIMOS 12 MESES (€/MWh)",f"=SUMPRODUCT(C{P0}:C{PLAST},E{P0}:E{PLAST})/SUM(E{P0}:E{PLAST})",hi=True)
prow(fr+2,"Peajes y ajustes (€/MWh)",None,yellow=CFG["peajes"])
prow(fr+3,"Cobertura de seguridad (€/MWh)",None,yellow=CFG["cobertura"])
prow(fr+4,"Precio fijo TEÓRICO hoy (€/MWh)",f"=C{fr+1}+C{fr+2}+C{fr+3}")
prow(fr+5,"Mejor fijo del catálogo — energía (€/MWh)",None,yellow=CFG["mejor_fijo"])
prow(fr+6,"VENTANA DE FIJACIÓN (teórico − catálogo)",f"=C{fr+4}-C{fr+5}",hi=True)
lect=wp.cell(row=fr+7,column=1,value=f'=IF(C{fr+6}>5,"VENTANA ABIERTA: quien fija hoy compra por debajo de donde repreciará el catálogo",IF(C{fr+6}<-5,"El catálogo está caro frente al mercado: negociar o esperar","Neutral: catálogo alineado con el mercado"))')
lect.font=Font(name=A,size=10,bold=True); lect.alignment=ll
wp.merge_cells(start_row=fr+7,start_column=1,end_row=fr+7,end_column=5)

# notas de lectura (columna G, junto a la tabla)
NOTES=[
("CÓMO SE LEE LA CURVA:",1),
("· Corto tenso: Sep-Dic 26 cotizan 97-103 €/MWh.",0),
("· Q2-27 a 45,85: el mercado espera primaveras hundidas (solar). Es la base de la jugada trimestral: cerrar Q1/Q3 y dejar Q2 indexado.",0),
("· Curva anual bajista: YR-27 66 → YR-29 53. El mercado espera normalización: cuidado con fijar 2-3 años a precios de hoy.",0),
("· Pesos del forward: 12 meses contiguos (por defecto Sep+Oct+Nov+Dic + Q1-27×3 + Q2-27×3 + Q3-27×2). Cuando cambie el mes, reajusta pesos.",0),
("· Fórmula sectorial del fijo (validada al euro): forward + ~28 peajes/ajustes + ~10 cobertura + fee.",0),
]
nr=5
for txt,b in NOTES:
    c=wp.cell(row=nr,column=7,value=txt); c.font=(lf if b else Font(name=A,size=9)); c.alignment=wl
    if b: c.fill=lg
    wp.row_dimensions[nr].height=26
    nr+=1

# ================================================== hojas de aterrizaje PQ
for nm,heads in [("PQ_Mensual",["Mes","P1","P2","P3"]),
                 ("PQ_Diario",["Fecha","Minimo","Media","Maximo"]),
                 ("PQ_Mensual6P",["Mes","P1","P2","P3","P4","P5","P6"]),
                 ("PQ_OMIP",["Clave","Instrumento","Precio"])]:
    ws=wb.create_sheet(nm)
    ws.sheet_view.showGridLines=False
    c=ws.cell(row=1,column=1,value=f"Zona de carga Power Query — consulta '{nm[3:] if nm!='PQ_Mensual6P' else 'OMIE_Mensual6P'}'. Cargar la tabla en A1. Vacía de fábrica: el panel usa los congelados.")
    c.font=sf
    # nota en fila 1; la tabla PQ aterriza en A1 y la sobreescribe — correcto

tmp=OUT+".tmp"
wb.save(tmp)
os.replace(tmp,OUT)
print("Guardado:",OUT)
print("Meses 2.0:",NFROZ,"congelados de",NROWS,"filas | 6P:",len(OM6),"| dias julio:",len(DAILY),"| OMIP:",len(OMIP_FROZ),"instrumentos")
fw=sum(p*w for _,_,p,w in OMIP_FROZ)/12
print("Forward congelado check: %.2f €/MWh -> teorico %.2f vs 109.9 -> ventana %.2f"%(fw,fw+38,fw+38-109.9))
